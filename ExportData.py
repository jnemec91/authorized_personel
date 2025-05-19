from Database import Database
from unidecode import unidecode
import openpyxl as opx

class ExportData:
    def __init__(self, db) -> None:
        if not isinstance(db, Database):
            self.db = Database(db)
        else:
            self.db = db


    def export_authorizations(self):
        export = opx.Workbook()
        sheet = export.active

        print('Exporting authorizations...')
        sheet.append(['Person Number', 'First Name', 'Last Name', 'Card Number', 'Email', 'Reader Number', 'Location Blueprint', 'Location Hospital', 'Location Name', 'ABI Location'])

        authorizations = self.db.get_authorizations()
        readers = self.db.get_readers()
        readers_with_authorizations = set([authorizations[1] for authorizations in authorizations])
        print('Readers_with_authorizations:', len(readers_with_authorizations))
        progress = 0
        complete = len(authorizations)
        for authorization in authorizations:
            progress += 1
            print(f"Progress: {'█'*(progress//(complete // 10))}{' '*(10-progress//(complete // 10))} {progress}/{complete}", end="\r")
            reader = self.db.select_reader(authorization[1])
            person = self.db.select_person(authorization[0])

            row = [person[0], person[1], person[2], person[3], person[4], reader[0], reader[1], reader[2], reader[3]]
            sheet.append(row)
        
        print('Checking for missing authorizations...')
        progress = 0
        complete = len(readers)
        count = 0
        for reader in readers:
            progress += 1
            print(f"Progress: {'█'*(progress//(complete // 10))}{' '*(10-progress//(complete // 10))} {progress}/{complete}", end="\r")
            if reader[0] not in readers_with_authorizations:
                row = [None, None, None, None, None, reader[0], reader[1], reader[2], reader[3], reader[4]]
                sheet.append(row)
                count += 1
        
        print(f'Found {count} missing authorizations.')
        export.save('output_documents/AuthorizationsOutput.xlsx')
        print('Export complete.')
        print('Output saved to output_documents/AuthorizationsOutput.xlsx')

    def export_readers(self):
        export = opx.Workbook()
        sheet = export.active

        for cell in sheet['A']:
            cell.number_format = 'Text'
        print('Setting cell format to text...')
        print('Exporting readers...')
        sheet.append(['Reader Number', 'Location Blueprint', 'Location Hospital', 'Location Name', 'ABI Location'])

        readers = self.db.get_readers()
        progress = 0
        complete = len(readers)
        for reader in readers:
            progress += 1
            print(f"Progress: {'█'*(progress//(complete // 10))}{' '*(10-progress//(complete // 10))} {progress}/{complete}", end="\r")
            sheet.append(reader)
        
        export.save('output_documents/ReadersOutput.xlsx')
        print('Export complete.')
        print('Output saved to output_documents/ReadersOutput.xlsx')
    
    def export_departments(self):
        # export in format dept_id (empty field), dept_name (number), dept_authorized (last_name+per_number), dept readers (reader_numbers separated by spaces)
        export = opx.Workbook()
        sheet = export.active

        print('Exporting departments...')
        sheet.append(['dept_id', 'dept_name', 'dept_authorized', 'dept_readers'])

        authorizations = self.db.get_authorizations()
        readers = self.db.get_readers()
        people = self.db.get_people()

        progress = 0
        complete = len(people)
        for person in people:
            progress += 1
            dept_readers = ''
            # remove czech special characters
            normalized_last_name = unidecode(person[2]).replace(' ', '').lower()

            dept_authorized = normalized_last_name + person[0]
            for authorization in authorizations:
                if authorization[0] == person[0]:
                    reader = self.db.select_reader(authorization[1])
                    dept_readers += reader[0] + ' '

            print(f"Progress: {'█'*(progress//(complete // 10))}{' '*(10-progress//(complete // 10))} {progress}/{complete}", end="\r") 
            sheet.append(['', person[2], dept_authorized, dept_readers])
        
        export.save('output_documents/DepartmentsOutput.xlsx')

if __name__ == '__main__':
    db = Database('test.db')
    export = ExportData(db)
    export.export_authorizations()
    export.export_readers() 
    export.export_departments()
